"""Parse uploaded PDFs and Excel files to extract financial data."""
import json
import re
from pathlib import Path
from typing import Dict, Any, Optional


def _looks_spaceless(text: str) -> bool:
    """Heuristic: pdfminer/pdfplumber sometimes drop the spaces between words
    on certain statement PDFs, gluing words into 'PropertyAddress',
    'AccountNumber', 'OutstandingPrincipal'. The tell-tale is an internal
    lower→upper case transition, which correctly spaced English text almost
    never has — so we use that rather than word length (legalese has plenty of
    legitimately long words like 'electronically')."""
    if not text:
        return True
    camel = re.findall(r'[a-z][A-Z]', text)
    return len(camel) >= 8


def _strip_md_tables(md: str) -> str:
    """Flatten MarkItDown's markdown tables into plain 'label value' lines so
    the field parser can read them. Drops separator rows and pipe scaffolding."""
    out = []
    for line in md.splitlines():
        s = line.strip()
        if re.fullmatch(r'\|?[\s|:\-]+\|?', s) and '-' in s:
            continue  # table separator row (| --- | --- |)
        if '|' in s:
            cells = [c.strip() for c in s.split('|') if c.strip()]
            if cells:
                out.append(' '.join(cells))
        else:
            out.append(line)
    return '\n'.join(out)


def markitdown_convert(filepath: str) -> str:
    """Convert a document to readable markdown with Microsoft MarkItDown."""
    from markitdown import MarkItDown
    return MarkItDown().convert(filepath).text_content or ""


def _pdfplumber_text(filepath: str, x_tolerance: float = 3) -> str:
    import pdfplumber
    text = ""
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text(x_tolerance=x_tolerance, y_tolerance=3)
            if page_text:
                text += page_text + "\n"
    return text


def _ocr_text(filepath: str) -> str:
    """Render each PDF page at 300 DPI and OCR with Tesseract.

    Last resort for scanned / image-only PDFs where no text layer exists.
    Requires: PyMuPDF (fitz), pytesseract, Pillow.
    """
    try:
        import fitz                 # PyMuPDF
        import pytesseract
        from PIL import Image
        import io
        doc = fitz.open(filepath)
        pages_text = []
        for page in doc:
            mat = fitz.Matrix(300 / 72, 300 / 72)   # 300 DPI
            pix = page.get_pixmap(matrix=mat)
            img = Image.open(io.BytesIO(pix.tobytes('png')))
            pages_text.append(pytesseract.image_to_string(img))
        return '\n'.join(pages_text)
    except Exception:
        return ''


def _clean_ocr_artifacts(text: str) -> str:
    """Repair common OCR artifacts in scanned financial documents."""
    # OCR sometimes inserts spaces inside dollar amounts: "$1 65,984.70" → "$165,984.70"
    text = re.sub(r'(\$\s*\d{1,3})\s+(\d{2,3},\d{3}(?:\.\d{1,2})?)', lambda m: m.group(1).replace(' ', '') + m.group(2), text)
    # "3.625 %" or "3.625  %" → "3.625%"
    text = re.sub(r'(\d+\.?\d*)\s{1,3}%', r'\1%', text)
    return text


def _camel_count(text: str) -> int:
    return len(re.findall(r'[a-z][A-Z]', text)) if text else 10 ** 9


def extract_pdf_text(filepath: str) -> str:
    """Extract readable plain text from a PDF for field parsing.

    Priority order:
      1. MarkItDown (pdfminer-based) — best for text PDFs with embedded fonts
      2. pdfplumber — recovers spacing when pdfminer glues words
      3. Tesseract OCR — last resort for scanned / image-only PDFs

    The winning candidate is whichever has the fewest camelCase run-togethers.
    OCR output is passed through _clean_ocr_artifacts() before returning.
    """
    candidates = []

    try:
        md_text = _strip_md_tables(markitdown_convert(filepath))
    except Exception:
        md_text = ""
    if md_text and len(md_text.strip()) > 30:
        if not _looks_spaceless(md_text):
            return md_text      # Fast path: clean text from MarkItDown
        candidates.append(md_text)

    # pdfplumber fallback with two tolerances
    for xt in (3, 1):
        try:
            t = _pdfplumber_text(filepath, x_tolerance=xt)
            if t and len(t.strip()) > 30:
                candidates.append(t)
        except Exception as e:
            if not candidates:
                candidates.append(f"Error extracting PDF: {e}")

    non_empty = [c for c in candidates if c and len(c.strip()) > 30]
    if non_empty:
        return min(non_empty, key=_camel_count)

    # Nothing worked — scanned / image-only PDF: fall back to OCR
    ocr = _ocr_text(filepath)
    if ocr and len(ocr.strip()) > 30:
        return _clean_ocr_artifacts(ocr)
    return '\n'.join(c for c in candidates if c) or ''


def extract_excel_data(filepath: str) -> Dict[str, Any]:
    """Extract data from Excel file."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            data[sheet_name] = rows
        return data
    except Exception as e:
        return {"error": str(e)}


def detect_category(text: str) -> str:
    """Guess the document category from its text content."""
    t = text.lower()
    # All closing / settlement documents → closing_statement
    # This covers: CFPB Closing Disclosure, ALTA Settlement Statement, HUD-1
    if re.search(r'closing\s+disclosure', t):
        return 'closing_statement'
    if re.search(r'alta\s+settlement\s+statement|hud-1\s+settlement\s+statement', t):
        return 'closing_statement'
    if re.search(r'sale\s+price\s+of\s+property.*settlement\s+date|settlement\s+date.*sale\s+price\s+of\s+property', t, re.DOTALL):
        return 'closing_statement'
    if re.search(r'loan\s+estimate|loan\s+disclosure', t):
        return 'loan_disclosure'
    # 1098/1099 forms must be matched before tax_return: their boilerplate
    # references "Form 1040 / Schedule A", which otherwise trips tax_return.
    if re.search(r'form\s*1098|mortgage\s+interest\s+statement|mortgage\s+interest\s+received\s+from', t):
        return '1098'
    if re.search(r'form\s*1099|\b1099\b', t):
        return '1099'
    if re.search(r'schedule\s*e|form\s*1040|tax\s+return', t):
        return 'tax_return'
    if re.search(r'mortgage\s+statement|escrow\s+balance|principal\s+balance', t):
        return 'mortgage_statement'
    if re.search(r'bank\s+statement|account\s+summary|checking\s+account|savings\s+account|beginning\s+balance|ending\s+balance', t):
        return 'bank_statement'
    if re.search(r'property\s+tax\s+statement|property\s+taxes?\s+for|assessed\s+value|parcel\s+#|full\s+cash\s+value|limited\s+value', t):
        return 'property_tax'
    return 'other'


def detect_period_type(data: Dict[str, Any]) -> str:
    """Detect document period from extracted dates.
    Returns: monthly, quarterly, half_yearly, yearly, other
    """
    start = data.get("period_start") or data.get("statement_date")
    end = data.get("period_end") or data.get("statement_date")
    if not start or not end:
        if data.get("payment_due_date") and data.get("statement_date"):
            return "monthly"
        if data.get("annual_rental_income") is not None or data.get("taxes_paid") is not None:
            return "yearly"
        if data.get("rents_received") is not None:
            return "yearly"
        return "other"

    from datetime import datetime
    for f in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            d1 = datetime.strptime(start, f)
            d2 = datetime.strptime(end, f)
            days = abs((d2 - d1).days)
            if days <= 35:
                return "monthly"
            if days <= 105:
                return "quarterly"
            if days <= 195:
                return "half_yearly"
            return "yearly"
        except (ValueError, TypeError):
            continue
    return "other"


def parse_currency(text: str) -> Optional[float]:
    """Extract first dollar amount from text."""
    match = re.search(r'\$?([\d,]+\.?\d*)', text.replace(',', ''))
    if match:
        try:
            return float(match.group(1).replace(',', ''))
        except ValueError:
            return None
    return None


def parse_percentage(text: str) -> Optional[float]:
    """Extract first percentage from text."""
    match = re.search(r'([\d.]+)\s*%', text)
    if match:
        try:
            return float(match.group(1))
        except ValueError:
            return None
    return None


# Column-2 labels that bleed into the address line on multi-column statements
_ADDRESS_NOISE = r'\s+(?=(?:Interest|Principal|Escrow|Late|Fees?|Total|Amount|Balance|Payment)\b)|\s*\$'

_STREET_SUFFIX = (
    r'(?:St(?:reet)?|Ave(?:nue)?|Dr(?:ive)?|Way|Ct|Court|Ln|Lane|Rd|Road|'
    r'Blvd|Boulevard|Pl(?:ace)?|Cir(?:cle)?|Ter(?:race)?|Pkwy|Parkway|'
    r'Hwy|Highway|Loop|Trail|Trl)'
)


def parse_property_address(text: str) -> Dict[str, Any]:
    """Extract the property address from document text.

    Handles single-line ("Property Address: 456 Oak Ave, Dallas, TX 75201"),
    two-line formats, and multi-column statement layouts where unrelated
    figures share the line ("Property address 6867 Syrah Dr Dublin,CA
    Interest $2,929.16" with the ZIP at the start of the next line).
    """
    # Tolerate the label with or without a space ("Property Address" /
    # "PropertyAddress"). Pick the occurrence whose value starts with a street
    # number — coupon stubs sometimes repeat the label with other text after it.
    data = {}
    line1 = line2 = ''
    for mm in re.finditer(r'property\s*address[:\s]*([^\n]*)\n?([^\n]*)?', text, re.IGNORECASE):
        cand1 = (mm.group(1) or '').strip()
        cand2 = (mm.group(2) or '').strip()
        if not cand1:
            cand1, cand2 = cand2, ''
        if re.match(r'^\d', cand1):  # value begins with a house number
            line1, line2 = cand1, cand2
            break
        if not line1:
            line1, line2 = cand1, cand2
    if not line1:
        return {}

    # Cut off second-column noise that got merged into the address line
    blob = re.split(_ADDRESS_NOISE, line1)[0].strip(' ,')
    if not blob:
        return {}

    # "City, ST 12345" continuation on the next line — match as a prefix so any
    # trailing table text ("...85258 Paid Last Paid Year to") is ignored.
    csz = re.match(
        r'^([A-Za-z][A-Za-z .\-]*?),?\s*([A-Z]{2})\s+(\d{5})(?:-\d{4})?\b', line2)
    csz_glued = re.match(r'^([A-Za-z]+?)([A-Z]{2})(\d{5})(?:-\d{4})?\b', line2)
    if csz:
        data['property_city'] = csz.group(1).strip()
        data['property_state'] = csz.group(2)
        data['property_zip'] = csz.group(3)
    elif csz_glued:  # run-together "SCOTTSDALEAZ85258"
        data['property_city'] = csz_glued.group(1).strip().title()
        data['property_state'] = csz_glued.group(2)
        data['property_zip'] = csz_glued.group(3)
    # ZIP wrapped to the start of the next line (multi-column layouts)
    elif (zm := re.match(r'^(\d{5})(?:-\d{4})?\b', line2)):
        data['property_zip'] = zm.group(1)

    parts = [p for p in (s.strip() for s in blob.split(',')) if p]
    if not parts:
        return {}
    street = parts[0]

    for tail in parts[1:]:
        sz = re.search(r'\b([A-Z]{2})\b\s*(\d{5})?(?:-\d{4})?', tail)
        if sz and (len(tail) <= 2 or sz.group(2)):
            data['property_state'] = sz.group(1)
            if sz.group(2):
                data['property_zip'] = sz.group(2)
            city = tail[:sz.start()].strip(' ,')
            if city:
                data['property_city'] = city
        else:
            data['property_city'] = tail

    # No comma between street and city: split after the street suffix
    if 'property_city' not in data:
        sm = re.match(r'^(\d+\s+.+?\b' + _STREET_SUFFIX + r'\.?)\s+([A-Za-z][A-Za-z .]*)$',
                      street, re.IGNORECASE)
        if sm:
            street = sm.group(1)
            data['property_city'] = sm.group(2).strip()

    # OCR fix: a scanned statement can render a leading house-number '1' as the
    # letter 'I'/'l' ("I 0308 E San Salvador Dr" → "10308 E San Salvador Dr").
    street = re.sub(r'^[Il]\s?(?=\d)', '1', street)

    data['property_address'] = street
    return data


def parse_borrowers_block(text: str) -> list:
    """Find borrower names from the mailing-address block: one or more
    all-caps name lines directly above 'STREET' / 'CITY ST 12345' lines."""
    lines = [l.strip() for l in text.splitlines()]
    for i, line in enumerate(lines[:-1]):
        if (re.match(r'^\d+\s+[A-Z0-9 .#\-]+$', line)
                and re.match(r'^[A-Z .]+\s[A-Z]{2}\s+\d{5}(-\d{4})?$', lines[i + 1])):
            names = []
            j = i - 1
            while (j >= 0 and len(names) < 4
                   and re.match(r"^[A-Z][A-Z .,'\-]+$", lines[j])
                   and not re.search(r'\d', lines[j])):
                names.insert(0, lines[j])
                j -= 1
            return names
    return []


def _find_amount(text: str, label_patterns: list) -> Optional[float]:
    """Find a dollar amount that follows any of the given label patterns.

    Tolerates 'Label: $1,234.56', 'Label $1,234.56' and table-style layouts
    where the value sits on the same line separated by whitespace or pipes.
    """
    for pat in label_patterns:
        m = re.search(pat + r'[^\d\n%]*?\$?\s*([\d,]+\.\d{1,2}|[\d,]{4,})\b',
                      text, re.IGNORECASE)
        if m:
            try:
                return float(m.group(1).replace(',', ''))
            except ValueError:
                continue
    return None


_MONTH_NAMES = {m.lower(): i for i, m in enumerate(
    ['January', 'February', 'March', 'April', 'May', 'June', 'July',
     'August', 'September', 'October', 'November', 'December'], 1)}
_MONTH_NAMES.update({m[:3].lower(): i for m, i in list(_MONTH_NAMES.items())})


def _find_date(text: str, label_patterns: list) -> Optional[str]:
    for pat in label_patterns:
        m = re.search(pat + r'[:|\s]*([\d]{1,2}/[\d]{1,4}(?:/[\d]{2,4})?|[\d\-]{6,10})',
                      text, re.IGNORECASE)
        if m:
            return m.group(1)
        # Long-form dates like "May 12, 2026" or "Sept 1 2025" -> MM/DD/YYYY
        m = re.search(
            pat + r'[:|\s]*([A-Za-z]{3,9})\.?\s+(\d{1,2}),?\s+(\d{4})',
            text, re.IGNORECASE)
        if m:
            mon = _MONTH_NAMES.get(m.group(1).lower()[:3])
            if mon:
                return f"{mon:02d}/{int(m.group(2)):02d}/{m.group(3)}"
    return None


def parse_mortgage_statement(text: str) -> Dict[str, Any]:
    """Extract structured fields from mortgage statement text.

    Targets the schema: property_address, borrowers, account_number,
    original/unpaid principal balance, interest rate, maturity date,
    amount due, principal/interest portions, escrow.
    """
    data = parse_property_address(text)

    borrowers_m = re.search(r'borrowers?[:|]\s*([^\n|]+)', text, re.IGNORECASE)
    if borrowers_m:
        names = [n.strip() for n in re.split(r';|&| and ', borrowers_m.group(1)) if n.strip()]
    else:
        names = parse_borrowers_block(text)
    for i, name in enumerate(names[:4], 1):
        data[f'borrower_{i}'] = name

    acct_m = re.search(
        r'(?:mortgage\s+)?(?:account|loan)\s+(?:number|no\.?|#)[:|\s]*([\dXx*\-]{4,})',
        text, re.IGNORECASE
    )
    if acct_m:
        data['account_number'] = acct_m.group(1)

    v = _find_amount(text, [
        r'original\s+principal\s+balance',
        r'original\s+loan\s+amount',
    ])
    if v is not None:
        data['original_amount'] = v

    v = _find_amount(text, [
        r'unpaid\s+principal\s+balance',
        r'outstanding\s+principal(?:\s+balance)?',
        r'current\s+principal\s+balance',
        r'unpaid\s+balance',
        r'outstanding\s+balance',
        r'principal\s+balance',
    ])
    if v is not None:
        data['current_balance'] = v

    # Rate may carry an ARM note, e.g. "Interest rate(Until 10/2032 pmt) 5.12500%"
    rate_m = re.search(r'interest\s+rate\s*(?:\(([^)]*)\))?[:|\s]*([\d.]+)\s*%',
                       text, re.IGNORECASE)
    if rate_m:
        data['interest_rate'] = float(rate_m.group(2))
        if rate_m.group(1):
            data['rate_note'] = rate_m.group(1).strip()
            if 'until' in data['rate_note'].lower():
                data['loan_type'] = 'ARM'

    v = _find_amount(text, [
        r'(?:monthly|regular)\s+payment(?:\s+amount)?',
        r'amount\s+due',
        r'automatic\s+payment\s+(?:amount|on)',
        r'total\s+payment',
    ])
    if v is None:
        # Statement layouts put the figure on the line below "Amount due"
        m = re.search(r'amount\s+due[^\n]*\n[^\n]*?\$\s*([\d,]+\.\d{2})', text, re.IGNORECASE)
        if m:
            v = float(m.group(1).replace(',', ''))
    if v is not None:
        data['monthly_payment'] = v

    v = _find_amount(text, [
        r'escrow\s*\(taxes\s*and\s*insurance\)',
        r'escrow\s+(?:amount|payment|portion)',
        r'escrow\s*\$',
    ])
    if v is not None:
        data['escrow_amount'] = v

    # Try to extract property tax separately from escrow breakdown
    v = _find_amount(text, [
        r'property\s+tax(?:es)?\s*(?:paid|amount|escrow)?',
        r'taxes?\s+paid\s+(?:in\s+)?escrow',
        r'taxes?\s+escrow',
        r'taxes:',  # bare "Taxes: $887.47" escrow line on Rocket statements
    ])
    if v is not None:
        data['property_tax_amount'] = v

    # "principal:" (colon) catches the bare "Principal: $531.46" payment-
    # breakdown label without matching "principal balance:" (no colon there).
    # The current payment's breakdown is listed before any YTD/life-of-loan
    # totals, so the first match is the right one.
    v = _find_amount(text, [
        r'principal\s+portion', r'principal\s+paid', r'principal\s*\$',
        r'principal:',
    ])
    if v is not None:
        data['principal_due'] = v

    v = _find_amount(text, [
        r'interest\s+portion', r'interest\s+paid', r'interest\s*\$',
        r'interest:',
    ])
    if v is not None:
        data['interest_due'] = v

    d = _find_date(text, [r'maturity\s+date'])
    if d:
        data['maturity_date'] = d

    d = _find_date(text, [r'statement\s+date'])
    if d:
        data['statement_date'] = d
        # Extract year from statement date (e.g. "01/15/2024" → 2024, "1/15/24" → 2024)
        parts = re.split(r'[/\-]', d)
        if len(parts) >= 3:
            y = parts[-1]
            if len(y) == 2:
                y = '20' + y
            try:
                data['statement_year'] = int(y)
            except ValueError:
                pass

    d = _find_date(text, [
        r'payment\s+due\s+date', r'next\s+due\s+date', r'due\s+date',
    ])
    if not d:
        m = re.search(r'payment\s+due[^\n]*\n\s*(\d{1,2}/\d{1,2}/\d{2,4})', text, re.IGNORECASE)
        if m:
            d = m.group(1)
    if d:
        data['payment_due_date'] = d

    return data


def parse_tax_return(text: str) -> Dict[str, Any]:
    """Extract rental income data from tax return."""
    data = {}

    # Schedule E rental income
    rental_income_m = re.search(
        r'(?:rents?\s+received|rental\s+income)[:\s]+\$?([\d,]+\.?\d*)',
        text, re.IGNORECASE
    )
    if rental_income_m:
        data['annual_rental_income'] = float(rental_income_m.group(1).replace(',', ''))

    # Depreciation
    depr_m = re.search(
        r'depreciation[:\s]+\$?([\d,]+\.?\d*)', text, re.IGNORECASE
    )
    if depr_m:
        data['annual_depreciation'] = float(depr_m.group(1).replace(',', ''))

    # Property taxes (Schedule E line 16)
    taxes_m = re.search(
        r'(?:taxes?|property\s+tax(?:es)?)[:\s]+\$?([\d,]+\.?\d*)', text, re.IGNORECASE
    )
    if taxes_m:
        data['taxes_paid'] = float(taxes_m.group(1).replace(',', ''))

    # Total expenses
    expenses_m = re.search(
        r'total\s+expenses[:\s]+\$?([\d,]+\.?\d*)', text, re.IGNORECASE
    )
    if expenses_m:
        data['total_expenses'] = float(expenses_m.group(1).replace(',', ''))

    return data


# ── Per-property tax return extraction (Schedule E + Schedule A) ─────────────────
# IRS forms print whole-dollar amounts with a trailing period ("3,585."), which
# lets us tell real values from line numbers/day counts. Schedule E Part I lays
# its three properties out in fixed columns A/B/C; MarkItDown flattens those
# columns, so we read word x-coordinates with pdfplumber and snap each value to
# its column. That recovers the right property even when a row is sparse (e.g.
# only one property has mortgage interest on a given line).

_TR_MONEY = re.compile(r'^\(?-?\$?[\d,]+\.\)?$')


def _tr_value(token: str) -> Optional[float]:
    neg = token.startswith('(') or token.startswith('-')
    t = token.strip('()$').replace(',', '').rstrip('.').lstrip('-')
    if not t:
        return None
    try:
        v = float(t)
    except ValueError:
        return None
    return -v if neg else v


def _tr_lines(words, tol: float = 3):
    """Group pdfplumber words into visual lines (same row within `tol` px)."""
    ws = sorted(words, key=lambda w: (w['top'], w['x0']))
    lines = []
    for w in ws:
        for ln in lines:
            if abs(ln['top'] - w['top']) <= tol:
                ln['w'].append(w)
                break
        else:
            lines.append({'top': w['top'], 'w': [w]})
    for ln in lines:
        ln['w'].sort(key=lambda x: x['x0'])
        ln['text'] = ' '.join(x['text'] for x in ln['w'])
    return sorted(lines, key=lambda l: l['top'])


def _tr_money_words(ln):
    out = []
    for w in ln['w']:
        if w['x0'] > 340 and _TR_MONEY.match(w['text']):
            v = _tr_value(w['text'])
            if v is not None:
                out.append((w['x0'], v))
    return out


def _parse_schedule_e_page(page) -> list:
    """Return the rental properties (address + key line items) on one
    Schedule E Part I page, columns recovered from word x-positions."""
    lines = _tr_lines(page.extract_words())

    # Column addresses listed under "1a Physical address of each property"
    addr = {}
    capture = False
    for ln in lines:
        if 'Physical address of each property' in ln['text']:
            capture = True
            continue
        if capture and re.match(r'^1\s*b\b|Type of Property', ln['text']):
            break
        if capture:
            m = re.match(r'^([ABC])\b\s+(.*)$', ln['text'])
            if m:
                addr[m.group(1)] = re.sub(r'^K\s+', '', m.group(2)).strip()
    if not addr:
        return []

    targets = {
        'rents': r'Rents received',
        'mortgage_interest': r'Mortgage interest paid to banks',
        'property_taxes': r'\b16 Taxes\b',
        'depreciation': r'Depreciation expense',
        'total_expenses': r'Total expenses',
    }
    found, anchor_xs = {}, []
    for ln in lines:
        for key, pat in targets.items():
            if key not in found and re.search(pat, ln['text']):
                mw = _tr_money_words(ln)
                found[key] = mw
                # Fully-populated rows define the column centers
                if key in ('rents', 'total_expenses', 'depreciation'):
                    anchor_xs += [x for x, _ in mw]

    # Cluster x-positions into column centers (left->right = A, B, C)
    anchor_xs.sort()
    clusters = []
    for x in anchor_xs:
        if clusters and x - clusters[-1][-1] <= 25:
            clusters[-1].append(x)
        else:
            clusters.append([x])
    centers = sorted(sum(c) / len(c) for c in clusters)
    cols = ['A', 'B', 'C'][:len(centers)]
    if not centers:
        return []

    def assign(mw):
        out = {}
        for x, v in mw:
            i = min(range(len(centers)), key=lambda i: abs(centers[i] - x))
            out[cols[i]] = v
        return out

    props = []
    for col, address in addr.items():
        vals = {k: assign(found.get(k, [])).get(col, 0.0) or 0.0 for k in targets}
        props.append({
            'address': address,
            'property_kind': 'rental',
            'rents_received': vals['rents'],
            'mortgage_interest': vals['mortgage_interest'],
            'property_taxes': vals['property_taxes'],
            'depreciation': vals['depreciation'],
            'total_expenses': vals['total_expenses'],
            'net_income': round(vals['rents'] - vals['total_expenses'], 2),
        })
    return props


def _parse_schedule_a_primary(text: str) -> Optional[dict]:
    """Primary-home mortgage interest (Sch A line 8a) and real estate taxes
    (line 5b). These are the homeowner's own residence, not rentals."""
    # Anchor on the actual Schedule A form header (unique) — not the many
    # incidental "Schedule A" references elsewhere in the return.
    m = re.search(r'SCHEDULE A\s+Itemized Deductions\s+OMB', text, re.IGNORECASE)
    if not m:
        return None
    region = text[m.start():m.start() + 8000]
    interest = re.search(r'\b8a\s+([\d,]+)\.', region)
    taxes = re.search(r'\b5b\s+([\d,]+)\.', region)
    if not interest and not taxes:
        return None
    return {
        'property_kind': 'primary',
        'mortgage_interest': float(interest.group(1).replace(',', '')) if interest else 0.0,
        'property_taxes': float(taxes.group(1).replace(',', '')) if taxes else 0.0,
        'rents_received': 0.0,
        'depreciation': 0.0,
        'total_expenses': 0.0,
        'net_income': 0.0,
    }


def parse_tax_return_properties(filepath: str) -> Dict[str, Any]:
    """Full per-property extraction from a 1040 tax return: every Schedule E
    rental plus the Schedule A primary residence. SSNs and taxpayer names are
    never read into the result — only property figures."""
    import pdfplumber

    full_text = extract_pdf_text(filepath)
    year_m = re.search(r'Schedule\s+E\s*\(Form\s*1040\)\s*(20\d{2})', full_text) \
        or re.search(r'Form\s*1040\s*\((20\d{2})\)', full_text) \
        or re.search(r'\b(20\d{2})\b', full_text)
    tax_year = int(year_m.group(1)) if year_m else None

    properties = []
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            txt = page.extract_text() or ''
            if 'Physical address of each property' in txt and 'Rents received' in txt:
                properties.extend(_parse_schedule_e_page(page))

    primary = _parse_schedule_a_primary(full_text)
    if primary:
        properties.append(primary)

    return {'tax_year': tax_year, 'properties': properties}


def _money_after(text: str, label: str, window: int = 120,
                 require_dollar: bool = True) -> Optional[float]:
    """Find the first dollar amount that follows a label.

    1098/1099 forms are multi-column; MarkItDown/pdfplumber flatten them so a
    box label and its value land on separate lines, sometimes with form noise
    ("Mortgage Interest Statement") in between. The window spans that gap.

    When *require_dollar* is False the amount may appear without a leading
    '$' sign — some servicer PDFs print Box 2 as a bare number on the next
    line, with the '$' appearing later next to the date in Box 3.
    """
    dollar = r'\$\s*' if require_dollar else r'(?:\$\s*)?'
    m = re.search(
        label + r'[\s\S]{0,%d}?%s([\d,]+\.\d{2})' % (window, dollar),
        text, re.IGNORECASE)
    if m:
        try:
            return float(m.group(1).replace(',', ''))
        except ValueError:
            return None
    return None


def _parse_1098_stacked(text: str) -> Dict[int, Any]:
    """Map a 'Box N:' label list onto the value list that follows it.

    Web year-end statements render every 1098 box label first and then every
    value in the same order (Box 1..Box 10, then $..$..date..$), so a per-label
    window grabs the wrong figure. Pair label and value by position instead.
    A 1099-INT block often follows, so the search is confined to the 1098
    section (between "Form 1098" and the next "Form 1099").
    """
    sec = text
    start = re.search(r'form\s*1098\b', text, re.IGNORECASE)
    if start:
        rest = text[start.end():]
        nxt = re.search(r'form\s*1099\b', rest, re.IGNORECASE)
        sec = rest[:nxt.start()] if nxt else rest
    labels = [int(n) for n in re.findall(r'box\s*(\d+)\s*:', sec, re.IGNORECASE)]
    if not labels:
        return {}
    last = list(re.finditer(r'box\s*\d+\s*:[^\n]*', sec, re.IGNORECASE))[-1].end()
    tokens = re.findall(r'(?:\$\s*)?[\d,]+\.\d{2}|\b\d{2}/\d{2}/\d{4}\b', sec[last:])
    out: Dict[int, Any] = {}
    for box, tok in zip(labels, tokens):
        if '/' in tok:
            out[box] = tok
        else:
            try:
                out[box] = float(tok.replace('$', '').replace(',', '').strip())
            except ValueError:
                pass
    return out


def parse_1098(text: str) -> Dict[str, Any]:
    """Extract data from a Form 1098 Mortgage Interest Statement.

    Maps the IRS boxes onto the app schema:
      Box 1  Mortgage interest received       -> mortgage_interest (annual)
      Box 2  Outstanding mortgage principal   -> current_balance
      Box 3  Mortgage origination date        -> origination_date
      Box 5  Mortgage insurance premiums      -> mortgage_insurance
      Box 6  Points paid                      -> points_paid
      Box 10 Other / Real estate taxes paid   -> property_tax_amount (annual)
    The property address comes from Box 8 (property securing the mortgage),
    not the borrower's mailing address.
    """
    data = {}

    # Calendar / tax year -> drives statement_date and the yearly period.
    # Layouts vary: printed IRS forms say "For calendar year YYYY"; web
    # snapshots say "Tax and Interest Information (YYYY)"; others only show
    # the year in the OMB header box next to "OMB No. 1545-1380".
    # On multi-column statements the year often lands several words after the
    # "For calendar year" label (column noise is interleaved), so allow a gap.
    # "Loan/Escrow Activity YYYY" sections on servicer statements are a
    # reliable fallback. The "(Rev. Month YYYY)" form-revision year always
    # precedes the calendar-year label, so a forward span won't pick it up.
    # Some servicer PDFs have character-level spacing ("c a l e n d a r   y e a r")
    # and the year can land 200+ chars away — widen the window and also try
    # a spacing-tolerant match.
    y = None
    for pat in (
        r'tax\s+and\s+interest\s+information\s*\((\d{4})\)',
        r'calendar\s+year[\s\S]{0,250}?\b(20\d{2})\b',
        r'c\s*\.?\s*a\s*\.?\s*l\s*\.?\s*e\s*\.?\s*n\s*\.?\s*d\s*\.?\s*a\s*\.?\s*r\s+\.?\s*y\s*\.?\s*e\s*\.?\s*a\s*\.?\s*r[\s\S]{0,250}?\b(20\d{2})\b',
        r'(?:loan|escrow)\s+activity\s+(20\d{2})',
        r'1545-1380[\s\S]{0,200}?\b(20\d{2})\b',
    ):
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            y = m.group(1)
            break
    if not y:
        # Last-resort fallback: take the first standalone 20xx year in the
        # first 30 lines of the document (before the instructions section).
        head = '\n'.join(text.splitlines()[:30])
        m = re.search(r'\b(20\d{2})\b', head)
        if m:
            y = m.group(1)
    if y:
        data['tax_year'] = y
        data['statement_year'] = int(y)
        data['statement_date'] = f"12/31/{y}"
        data['period_start'] = f"01/01/{y}"
        data['period_end'] = f"12/31/{y}"

    # Box amounts. "Box N:" colon layouts stack labels then values, so map
    # them positionally; printed forms keep each value near its label.
    boxes = _parse_1098_stacked(text) if re.search(r'box\s*\d+\s*:', text, re.IGNORECASE) else {}
    if boxes:
        if boxes.get(1) is not None:
            data['mortgage_interest'] = boxes[1]
        if boxes.get(2) is not None:
            data['current_balance'] = boxes[2]
        if boxes.get(5):
            data['mortgage_insurance'] = boxes[5]
        if boxes.get(6):
            data['points_paid'] = boxes[6]
        if boxes.get(10) is not None:
            data['property_tax_amount'] = boxes[10]
        if isinstance(boxes.get(3), str):
            data['origination_date'] = boxes[3]
    else:
        # pdfplumber sometimes splits "Mortgage" into "M ortgage" on these
        # forms, so anchor on the "interest received from" portion, which
        # survives the spacing glitch.
        v = _money_after(text, r'interest\s+received\s+from', 90)
        if v is not None:
            data['mortgage_interest'] = v
        v = _money_after(text, r'outstanding\s+mortgage\s+principal', 130)
        if v is not None:
            data['current_balance'] = v
        else:
            # Fallback: some servicer PDFs print Box 2 as a bare number
            # without a '$' sign (the '$' lands on the next line next to
            # the Box 3 date).  Allow the amount without the dollar symbol.
            v = _money_after(text, r'outstanding\s+mortgage\s+principal', 130,
                             require_dollar=False)
            if v is not None:
                data['current_balance'] = v
        v = _money_after(text, r'mortgage\s+insurance\s+premiums?', 30)
        if v is not None:
            data['mortgage_insurance'] = v
        v = _money_after(text, r'points\s+paid', 40)
        if v is not None:
            data['points_paid'] = v
        v = _money_after(text, r'real\s+estate\s+taxes\s+paid', 40)
        if v is None:
            v = _money_after(text, r'\b10\s+other\b', 30)
        if v is not None:
            data['property_tax_amount'] = v

        m = re.search(r'mortgage\s+origination\s+date[\s\S]{0,80}?(\d{2}/\d{2}/\d{4})', text, re.IGNORECASE)
        if m:
            data['origination_date'] = m.group(1)

    m = re.search(
        r'account\s+number(?:\s*\(see\s+instructions\))?[\s:]*\n+\s*([\dXx\*\-]{5,})',
        text, re.IGNORECASE)
    if m:
        data['account_number'] = m.group(1).strip()

    m = re.search(
        r"recipient'?s/lender'?s\s+name[\s\S]*?(?:telephone no\.?|postal code[^\n]*)\s*\n\s*([^\n]+)",
        text, re.IGNORECASE)
    if m:
        name = m.group(1).strip()
        # Reject lines that are really wrapped form boilerplate, not a name
        boilerplate = ('telephone', 'limits based', 'deductible', 'postal',
                       'caution', 'secured property', '1098')
        if name and not any(w in name.lower() for w in boilerplate):
            data['lender_name'] = name

    # Box 8 - address/description of the property securing the mortgage
    m = re.search(
        r'address or description of property securing mortgage\s*\n+\s*([^\n]+)\n+\s*([^\n]+)',
        text, re.IGNORECASE)
    if m:
        data['property_address'] = re.sub(r'\s{2,}', ' ', m.group(1).strip())
        csz = re.match(r'^(.*?)[,]?\s+([A-Z]{2})\s+(\d{5})(?:-\d{4})?\b', m.group(2).strip())
        if csz:
            data['property_city'] = csz.group(1).strip().title()
            data['property_state'] = csz.group(2)
            data['property_zip'] = csz.group(3)

    # Borrower name(s) - capture lines after the wrapped label, before the street
    m = re.search(
        r"payer'?s/borrower'?s\s+name[\s\S]*?postal code\s*\n([\s\S]*?)\n\s*\d",
        text, re.IGNORECASE)
    if m:
        names = [l.strip() for l in m.group(1).splitlines()
                 if l.strip() and not re.search(r'\d', l)]
        for i, nm in enumerate(names[:4], 1):
            data[f'borrower_{i}'] = nm

    return data


def parse_1099(text: str) -> Dict[str, Any]:
    """Extract data from 1099 form."""
    data = {}
    # Box 1 - rents
    box1_m = re.search(r'(?:box\s*1|rents)[:\s]+\$?([\d,]+\.?\d*)', text, re.IGNORECASE)
    if box1_m:
        data['rents_received'] = float(box1_m.group(1).replace(',', ''))
    return data


def parse_bank_statement(text: str) -> Dict[str, Any]:
    """Extract data from a bank statement."""
    data = parse_property_address(text)

    # Bank name
    bank_m = re.search(r'([A-Za-z\s]+(?:bank|credit union|financial|savings))',
                       text, re.IGNORECASE)
    if bank_m:
        data['bank_name'] = bank_m.group(1).strip()

    # Account number
    acct_m = re.search(
        r'(?:account|acct)\s+(?:number|no\.?|#)[:|\s]*([\dXx*\-]{4,})',
        text, re.IGNORECASE
    )
    if acct_m:
        data['account_number'] = acct_m.group(1)

    # Statement period
    period_m = re.search(
        r'(?:statement\s+)?(?:period|for\s+the\s+month|for\s+period)[\s:]+\
([A-Za-z]+\s+\d{1,2},?\s*\d{4})\s*(?:through|to|-)\s*([A-Za-z]+\s+\d{1,2},?\s*\d{4})',
        text, re.IGNORECASE
    )
    if not period_m:
        period_m = re.search(
            r'(?:statement\s+)?(?:period|for\s+the\s+month|for\s+period)[\s:]+\
(\d{1,2}/\d{1,2}/\d{2,4})\s*(?:through|to|-)\s*(\d{1,2}/\d{1,2}/\d{2,4})',
            text, re.IGNORECASE
        )
    if period_m:
        data['period_start'] = period_m.group(1)
        data['period_end'] = period_m.group(2)

    # Beginning balance
    v = _find_amount(text, [
        r'beginning\s+balance', r'starting\s+balance', r'opening\s+balance',
    ])
    if v is not None:
        data['beginning_balance'] = v

    # Ending balance
    v = _find_amount(text, [
        r'ending\s+balance', r'closing\s+balance',
    ])
    if v is not None:
        data['ending_balance'] = v

    # Total deposits / credits
    v = _find_amount(text, [
        r'total\s+deposits?', r'total\s+credits?', r'deposits?\s+and\s+credits?',
    ])
    if v is not None:
        data['total_deposits'] = v

    # Total withdrawals / debits
    v = _find_amount(text, [
        r'total\s+withdrawals?', r'total\s+debits?', r'withdrawals?\s+and\s+debits?',
    ])
    if v is not None:
        data['total_withdrawals'] = v

    # No statement_date — use period_end
    if not data.get('statement_date') and data.get('period_end'):
        data['statement_date'] = data['period_end']

    return data


def parse_property_tax(text: str) -> Dict[str, Any]:
    """Extract property tax details from annual/half-yearly tax statements."""
    data = parse_property_address(text)

    # Tax year. The county's own "TAX YEAR: YYYY" stub is authoritative; a
    # fiscal-range title like "2023 - 2024 PROPERTY TAX BILL" otherwise gets
    # mis-read as the later year, colliding bills from adjacent years.
    year_m = re.search(r'tax\s+year[:\s]+(\d{4})', text, re.IGNORECASE)
    if not year_m:
        # Fiscal-range title: use the starting year ("2023 - 2024" -> 2023).
        year_m = re.search(
            r'(\d{4})\s*[-–]\s*\d{4}\s+property\s+tax', text, re.IGNORECASE)
    if not year_m:
        year_m = re.search(r'(\d{4})\s+property\s+tax', text, re.IGNORECASE)
    if not year_m:
        year_m = re.search(r'property\s+tax\s+\w+\s+(\d{4})', text, re.IGNORECASE)
    if not year_m:
        year_m = re.search(r'(\d{4})\s+assessed\s+taxes', text, re.IGNORECASE)
    if year_m:
        data['tax_year'] = year_m.group(1)
        data['statement_date'] = f"01/01/{year_m.group(1)}"
        data['statement_year'] = int(year_m.group(1))
        data['period_start'] = f"01/01/{year_m.group(1)}"
        data['period_end'] = f"12/31/{year_m.group(1)}"

    # Total taxes
    v = _find_amount(text, [
        r'total\s+\d{4}\s+property\s+taxes?',
        r'total\s+assessed\s+taxes?',
        r'your\s+total\s+\d{4}\s+property\s+taxes?',
        r'your\s+total\s+property\s+taxes?',
    ])
    # County installment bills (e.g. San Joaquin CA) show no single "total"
    # line — the annual tax is the sum of the 1st and 2nd installments.
    if v is None:
        installments = re.findall(
            r'(?:1st|2nd|first|second)\s+installment\s*\$?\s*([\d,]+\.\d{2})',
            text, re.IGNORECASE)
        if installments:
            v = round(sum(float(x.replace(',', '')) for x in installments), 2)
    if v is not None:
        data['taxes_paid'] = v
        data['property_tax_amount'] = v

    # Parcel / APN
    parcel_m = re.search(r'parcel[#\s]+([\d\-]+)', text, re.IGNORECASE)
    if parcel_m:
        data['parcel_number'] = parcel_m.group(1)

    # Assessed values
    v = _find_amount(text, [r'limited\s+value'])
    if v is not None:
        data['limited_value'] = v
    v = _find_amount(text, [r'full\s+cash\s+value', r'secondary\s+full\s+cash'])
    if v is not None:
        data['full_cash_value'] = v

    # Half-year amounts from payment stubs
    halves = re.findall(r'(?:first|second)\s+half[^$]*\$?\s*([\d,]+\.\d{2})', text, re.IGNORECASE)
    if halves:
        data['first_half'] = float(halves[0].replace(',', ''))
        if len(halves) > 1:
            data['second_half'] = float(halves[1].replace(',', ''))

    # Previous year total for comparison
    v = _find_amount(text, [r'previous\s+year\s+total'])
    if v is not None:
        data['previous_year_taxes'] = v

    return data


def parse_loan_disclosure(text: str) -> Dict[str, Any]:
    """Extract loan terms from Loan Estimate / Closing Disclosure."""
    data = parse_property_address(text)

    # Loan amount
    loan_m = re.search(
        r'loan\s+amount[:\s]+\$?([\d,]+\.?\d*)', text, re.IGNORECASE
    )
    if loan_m:
        data['original_amount'] = float(loan_m.group(1).replace(',', ''))

    # Rate
    rate_m = re.search(r'interest\s+rate[:\s]+([\d.]+)\s*%', text, re.IGNORECASE)
    if rate_m:
        data['interest_rate'] = float(rate_m.group(1))

    # Loan term
    term_m = re.search(r'loan\s+term[:\s]+(\d+)\s*years?', text, re.IGNORECASE)
    if term_m:
        data['loan_term_years'] = int(term_m.group(1))

    # Loan type
    if re.search(r'\bARM\b|\badjustable\b', text, re.IGNORECASE):
        data['loan_type'] = 'ARM'
    elif re.search(r'\bfixed\b', text, re.IGNORECASE):
        data['loan_type'] = 'FIXED'

    # Monthly payment
    payment_m = re.search(
        r'(?:principal\s*&\s*interest|monthly\s+payment)[:\s]+\$?([\d,]+\.?\d*)',
        text, re.IGNORECASE
    )
    if payment_m:
        data['monthly_payment'] = float(payment_m.group(1).replace(',', ''))

    return data


def _normalize_date(s: str) -> Optional[str]:
    """Convert human-readable date strings to YYYY-MM-DD; return as-is if not parseable."""
    from datetime import datetime
    for fmt in ("%B %d, %Y", "%B %d %Y", "%b %d, %Y", "%b %d %Y",
                "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s.strip()


def _parse_amount(s: str) -> Optional[float]:
    """Parse a dollar amount string with or without cents."""
    try:
        return float(re.sub(r'[,$\s]', '', s))
    except (ValueError, TypeError):
        return None


def parse_closing_statement(text: str) -> Dict[str, Any]:
    """Extract fields from any closing / settlement document.

    Handles three formats generically:
      - CFPB Closing Disclosure (standardised since 2015)
      - ALTA Settlement Statement (title-company form)
      - HUD-1 Settlement Statement (legacy form)

    All three share the same output schema; field names follow the app's
    Loan / Property model conventions so _apply_extracted() works unchanged.
    """
    data = parse_property_address(text)  # general address extraction first

    # ── Property address — CFPB "Security Interest" section is most reliable ──
    # "You are granting a security interest in\n10575 East Mission Lane, Scottsdale, AZ 85258"
    m = re.search(
        r'security\s+interest\s+in\s*\n?\s*'
        r'(\d+[^\n,]+)(?:,|\n\s*)([^,\n]+),\s*([A-Z]{2})\s+(\d{5})',
        text, re.IGNORECASE
    )
    if m:
        data['property_address'] = m.group(1).strip()
        data['property_city']    = m.group(2).strip()
        data['property_state']   = m.group(3)
        data['property_zip']     = m.group(4)

    # ── Escrow / file number ──────────────────────────────────────────────────
    m = re.search(r'file\s+(?:no\.?|#)\s*(?:/\s*escrow\s+no\.?)?\s*([A-Z0-9][-A-Z0-9/]{2,40})',
                  text, re.IGNORECASE)
    if m:
        data['escrow_number'] = m.group(1).strip()

    # ── Loan ID / reference number ────────────────────────────────────────────
    m = re.search(r'loan\s*id\s*#?\s*(\d{6,20})', text, re.IGNORECASE)
    if m:
        data['loan_id'] = m.group(1)

    # ── Borrower ─────────────────────────────────────────────────────────────
    # ALTA: "Borrower: Full Name" at start of line
    # CFPB: "... Borrower Pavani Donepudi LoanTerm 30 years" (mid-line Transaction Info)
    # Non-greedy {1,3}? prevents consuming the following "LoanTerm" keyword
    for pat in [
        r'^borrower\s*:\s*(.+)',                                             # ALTA colon
        r'\bBorrower\s+([A-Z][a-z]+(?:\s+[A-Za-z]+){1,3}?)(?=\s+(?:Loan|Purpose|Product|\d))',  # CFPB inline
        r'^Borrower\s+([A-Z][a-z]+(?:\s+[A-Za-z]+){1,3})\s*$',             # CFPB addendum line
    ]:
        m = re.search(pat, text, re.MULTILINE)
        if m:
            candidate = (m.group(1) or '').strip()
            if len(candidate) > 3 and not candidate.startswith('-'):
                data['borrower_1'] = candidate
                break

    # ── Seller ────────────────────────────────────────────────────────────────
    # ALTA: "Seller: Name" at line start   CFPB: "... Seller = Gerald I Blackman ..."
    for pat in [
        r'^seller\s*:\s*(.+)',                                               # ALTA colon
        r'\bSeller\b\s*=\s*([A-Za-z][^\n]+?)(?=\s*\n|\s{3,})',             # CFPB "= Name"
    ]:
        m = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if m:
            candidate = m.group(1).strip()
            if candidate and not candidate.lower().startswith(('credit', 'paid', 'due')):
                data['seller_name'] = candidate
                break

    # ── Lender ────────────────────────────────────────────────────────────────
    # ALTA: "Lender: Company Name" at line start
    # CFPB: "Lender Better Mortgage Corporation LoanID# ..."
    # CFPB variant: "Lender = DHi Mortgage Company, Ltd., LP. LoanID# ..."
    for pat in [
        r'^lender\s*:\s*(.+)',                                               # ALTA colon
        r'\bLender\s*=\s*([A-Za-z][^\n]+?)(?:\s+Loan(?:ID|Type|Term)|$)',   # "= Name" then LoanID or EOL
        r'\bLender\s+([A-Za-z][A-Za-z\s,.-]+?)(?=\s+Loan(?:ID|Type|Term)|\n)',  # CFPB before LoanID/etc.
    ]:
        m = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if m:
            candidate = m.group(1).strip().rstrip('.')
            # Reject single generic words like "Credits", "Charges", "Name"
            if len(candidate) > 8 and (' ' in candidate or ',' in candidate):
                data['lender_name'] = candidate
                break

    # ── Closing / Settlement date ─────────────────────────────────────────────
    # CFPB: "Closing Date 5/27/2021"   ALTA: "Settlement Date : September 20, 2024"
    date_pattern = r'(?:closing|settlement)\s+date\s*:?\s*([A-Za-z]+\s+\d{1,2},?\s*\d{4}|\d{1,2}/\d{1,2}/\d{2,4})'
    m = re.search(date_pattern, text, re.IGNORECASE)
    if m:
        normed = _normalize_date(m.group(1).strip())
        data['settlement_date']  = normed
        data['purchase_date']    = normed
        data['origination_date'] = normed

    # ── Sale / Purchase price ─────────────────────────────────────────────────
    # CFPB: "Sale Price $655,000"   ALTA: "Sale Price of Property 675,200.00"
    m = re.search(
        r'sale\s+price(?:\s+of\s+(?:property|any\s+personal\s+property))?\s+\$?\s*([\d,]+(?:\.\d{1,2})?)',
        text, re.IGNORECASE
    )
    if m:
        v = _parse_amount(m.group(1))
        if v:
            data['purchase_price'] = v

    # ── Loan amount ───────────────────────────────────────────────────────────
    # Inline (ALTA): "Loan Amount $468,750.00" on the same line
    m = re.search(r'^loan\s+amount\s+\$?\s*([\d,]+(?:\.\d{1,2})?)', text, re.IGNORECASE | re.MULTILINE)
    if m:
        v = _parse_amount(m.group(1))
        if v:
            data['original_amount'] = v
    # CFPB two-column layout: "Loan Amount" on its own line; value appears as a
    # standalone "$NNN,NNN" line several lines later (right-column flush).
    if not data.get('original_amount'):
        m = re.search(r'^loan\s+amount\s*$', text, re.IGNORECASE | re.MULTILINE)
        if m:
            rest = text[m.end():]
            val_m = re.search(r'^\$?\s*(\d{1,3}(?:,\d{3})+(?:\.\d{1,2})?)\s*$', rest, re.MULTILINE)
            if val_m:
                v = _parse_amount(val_m.group(1))
                if v and v > 10000:
                    data['original_amount'] = v

    # ── Down payment ──────────────────────────────────────────────────────────
    # CFPB: "Down Payment/Funds from Borrower $164,000.00"
    m = re.search(r'down\s+payment[^$\n]*\$?\s*([\d,]+(?:\.\d{1,2})?)', text, re.IGNORECASE)
    if m:
        v = _parse_amount(m.group(1))
        if v:
            data['down_payment'] = v
    # ALTA: computed from sale price − loan amount
    elif data.get('purchase_price') and data.get('original_amount'):
        data['down_payment'] = round(data['purchase_price'] - data['original_amount'], 2)

    # ── Deposit / earnest money (ALTA) ────────────────────────────────────────
    m = re.search(r'^deposit\s+\$?\s*([\d,]+(?:\.\d{1,2})?)', text, re.IGNORECASE | re.MULTILINE)
    if m:
        v = _parse_amount(m.group(1))
        if v:
            data['deposit'] = v

    # ── Seller credit ─────────────────────────────────────────────────────────
    # Require a dollar sign OR comma-formatted amount to avoid matching bare line
    # numbers like "05" in "05 Seller Credit 05 Payoff of..."
    m = re.search(r'seller\s+credit\s+\$\s*([\d,]+(?:\.\d{1,2})?)', text, re.IGNORECASE)
    if m:
        v = _parse_amount(m.group(1))
        if v:
            data['seller_credit'] = v

    # ── Cash to close ─────────────────────────────────────────────────────────
    # CFPB page 3 "Summary of Transactions" has the definitive number
    m = re.search(r'cash\s+to\s+close\s+\$?\s*([\d,]+(?:\.\d{1,2})?)', text, re.IGNORECASE)
    if m:
        v = _parse_amount(m.group(1))
        if v:
            data['cash_to_close'] = v

    # ── Interest rate ─────────────────────────────────────────────────────────
    # CFPB inline: "Interest Rate 3.625%" on same line
    m = re.search(r'interest\s+rate\s+([\d.]+)%', text, re.IGNORECASE)
    if m:
        data['interest_rate'] = float(m.group(1))
    # CFPB two-column split: "Interest Rate" on its own line, "7.625%" several lines later
    if not data.get('interest_rate'):
        m = re.search(r'^interest\s+rate\s*$', text, re.IGNORECASE | re.MULTILINE)
        if m:
            rest = text[m.end():]
            val_m = re.search(r'^([\d.]+)\s*%\s*$', rest, re.MULTILINE)
            if val_m:
                v = float(val_m.group(1))
                if 0.1 < v < 30:
                    data['interest_rate'] = v
    if not data.get('interest_rate'):
        # ALTA fallback: derive from prepaid daily interest
        m = re.search(r'prepaid\s+interest\s+\$?([\d.]+)\s+per\s+day', text, re.IGNORECASE)
        if m and data.get('original_amount'):
            daily = float(m.group(1))
            data['interest_rate'] = round(daily * 365 / data['original_amount'] * 100, 2)

    # ── Monthly P&I payment ───────────────────────────────────────────────────
    # CFPB inline: "Principal & Interest $2,239.21" on same line
    m = re.search(
        r'principal\s*[&+]\s*interest\s+\$?\s*([\d,]+\.\d{2})',
        text, re.IGNORECASE
    )
    if m:
        data['monthly_payment'] = _parse_amount(m.group(1))
    # CFPB two-column split: "Principal & Interest" on its own line, value several lines later
    if not data.get('monthly_payment'):
        m = re.search(r'^(?:monthly\s+)?principal\s*&\s*interest\s*$', text, re.IGNORECASE | re.MULTILINE)
        if m:
            rest = text[m.end():]
            val_m = re.search(r'^\$?\s*(\d{1,3}(?:,\d{3})*\.\d{2})\s*$', rest, re.MULTILINE)
            if val_m:
                v = _parse_amount(val_m.group(1))
                if v and v > 100:
                    data['monthly_payment'] = v

    # ── Loan term ─────────────────────────────────────────────────────────────
    m = re.search(r'loan\s*term\s+(\d+)\s*years?', text, re.IGNORECASE)
    if m:
        data['loan_term_years'] = int(m.group(1))

    # ── Loan type ─────────────────────────────────────────────────────────────
    if re.search(r'\bARM\b|\badjustable[\s-]rate\b', text, re.IGNORECASE):
        data['loan_type'] = 'ARM'
    elif re.search(r'fixed[\s-]rate|product\s+fixed|\bconventional\b', text, re.IGNORECASE):
        data['loan_type'] = 'FIXED'

    # ── Annual Property Tax ───────────────────────────────────────────────────
    # CFPB escrow section: "Property Taxes $682.29 per month for N mo."
    m = re.search(r'property\s+taxes?\s+\$?([\d,]+\.\d{2})\s+per\s+month', text, re.IGNORECASE)
    if m:
        data['annual_property_tax'] = round(float(m.group(1).replace(',', '')) * 12, 2)
    # ALTA impound: "Property Taxes to … N Months at $791.39/month"
    if not data.get('annual_property_tax'):
        m = re.search(r'property\s+taxes?\s+to[^\n]*?\$?([\d,]+\.\d{2})\s*/\s*month',
                      text, re.IGNORECASE)
        if m:
            data['annual_property_tax'] = round(float(m.group(1).replace(',', '')) * 12, 2)

    # ── Annual Homeowner's Insurance ──────────────────────────────────────────
    # CFPB: "Homeowner's Insurance Premium (12 mo.) to Safeco $1,113.00"
    m = re.search(
        r"homeowner.{0,3}s\s+insurance\s+premium\s*\(12\s+mo\.[^)]*\)[^\n]*?\$\s*([\d,]+\.\d{2})",
        text, re.IGNORECASE
    )
    if m:
        data['annual_insurance'] = _parse_amount(m.group(1))
    # CFPB escrow detail: "Homeowner's Insurance $92.75 per month for N mo."
    # Handle OCR artifacts like "—_" between "Insurance" and "$"
    if not data.get('annual_insurance'):
        m = re.search(r"homeowner.{0,3}s\s+insurance[^\n$]*?\$\s*([\d,]+(?:\.\d{2})?)\s+per\s+month",
                      text, re.IGNORECASE)
        if m:
            data['annual_insurance'] = round(float(m.group(1).replace(',', '')) * 12, 2)
    # ALTA: "Homeowner's Insurance Premium to … 12 months NNN.NN"
    if not data.get('annual_insurance'):
        m = re.search(r"homeowner'?s?\s+insurance\s+premium[^\n]+12\s+months?\s+([\d,]+\.\d{2})",
                      text, re.IGNORECASE)
        if m:
            data['annual_insurance'] = _parse_amount(m.group(1))
    # ALTA monthly: "Homeowner's Insurance to … $59.85/month"
    if not data.get('annual_insurance'):
        m = re.search(r"homeowner'?s?\s+insurance\s+to[^\n]*?\$?([\d,]+\.\d{2})\s*/\s*month",
                      text, re.IGNORECASE)
        if m:
            data['annual_insurance'] = round(float(m.group(1).replace(',', '')) * 12, 2)

    # ── HOA (non-escrowed recurring cost) ─────────────────────────────────────
    # CFPB: "Non-Escrowed Property Costs over Year 1 $394.92" (HOA Dues)
    m = re.search(r'non[\s-]escrowed\s+property\s+costs[^\n]*?([\d,]+\.\d{2})', text, re.IGNORECASE)
    if m:
        data['hoa_annual'] = _parse_amount(m.group(1))

    # ── Monthly escrow payment ────────────────────────────────────────────────
    # CFPB: "Estimated Escrow + 775.04"  or  "Monthly $775.04"
    m = re.search(r'estimated\s+escrow\s+\+\s*([\d,]+\.\d{2})', text, re.IGNORECASE)
    if m:
        data['escrow_monthly'] = _parse_amount(m.group(1))

    # ── APR (informational) ───────────────────────────────────────────────────
    m = re.search(r'annual\s+percentage\s+rate\s*\(apr\)[^\n]*?([\d.]+)%', text, re.IGNORECASE)
    if m:
        data['apr'] = float(m.group(1))

    return data


CATEGORY_TITLES = {
    'mortgage_statement': 'Mortgage Statement',
    'tax_return': 'Tax Return',
    '1098': 'Form 1098 - Mortgage Interest Statement',
    '1099': 'Form 1099 Year-End Statement',
    'loan_disclosure': 'Loan Disclosure',
    'closing_statement': 'ALTA Settlement Statement / Closing Statement',
    'bank_statement': 'Bank Statement',
    'property_tax': 'Property Tax Statement',
    'other': 'Document',
}


def _md_table(rows: list) -> list:
    lines = ['| Field | Value |', '|---|---|']
    for label, value in rows:
        lines.append(f'| {label} | {value} |')
    lines.append('')
    return lines


def _fmt_value(key: str, v: Any) -> str:
    if isinstance(v, float):
        if 'rate' in key:
            return f'{v}%'
        return f'${v:,.2f}'
    return str(v)


def to_markdown(category: str, data: Dict[str, Any]) -> str:
    """Render extracted fields as a structured markdown document.

    Mortgage statements get the sectioned layout (Property Information,
    Payment Information, Payment Breakdown, Extraction Schema); other
    categories get a single field table.
    """
    title = CATEGORY_TITLES.get(category, 'Document')
    lines = [f'# {title} - Structured Fields', '']

    def section(source, name, fields):
        rows = [
            (label, _fmt_value(k, source[k]))
            for k, label in fields if source.get(k) is not None
        ]
        if rows:
            lines.append(f'## {name}')
            lines.append('')
            lines.extend(_md_table(rows))

    if category == 'mortgage_statement':
        display = dict(data)
        full_addr = data.get('property_address', '')
        if data.get('property_city'):
            full_addr += f", {data['property_city']}"
        if data.get('property_state'):
            full_addr += f", {data['property_state']} {data.get('property_zip', '')}".rstrip()
        if full_addr:
            display['property_address'] = full_addr
        borrowers = '; '.join(
            data[k] for k in ('borrower_1', 'borrower_2', 'borrower_3', 'borrower_4')
            if data.get(k)
        )
        if borrowers:
            display['borrowers'] = borrowers

        section(display, 'Property Information', [
            ('property_address', 'Property Address'),
            ('borrowers', 'Borrowers'),
            ('account_number', 'Mortgage Account Number'),
            ('original_amount', 'Original Principal Balance'),
            ('current_balance', 'Unpaid Principal Balance'),
            ('interest_rate', 'Interest Rate'),
            ('maturity_date', 'Maturity Date'),
        ])
        section(display, 'Payment Information', [
            ('statement_date', 'Statement Date'),
            ('payment_due_date', 'Payment Due Date'),
            ('monthly_payment', 'Amount Due'),
            ('escrow_amount', 'Escrow Amount'),
        ])
        section(display, 'Current Payment Breakdown', [
            ('principal_due', 'Principal Portion'),
            ('interest_due', 'Interest Portion'),
        ])

    if category == 'closing_statement':
        display = dict(data)
        full_addr = data.get('property_address', '')
        if data.get('property_city'):
            full_addr += f", {data['property_city']}"
        if data.get('property_state'):
            full_addr += f", {data['property_state']} {data.get('property_zip', '')}".rstrip()
        if full_addr:
            display['property_address'] = full_addr

        section(display, 'Transaction Details', [
            ('escrow_number', 'Escrow / File Number'),
            ('settlement_date', 'Settlement Date'),
            ('property_address', 'Property Address'),
            ('borrower_1', 'Borrower(s)'),
            ('seller_name', 'Seller'),
            ('lender_name', 'Lender'),
        ])
        section(display, 'Purchase Financials', [
            ('purchase_price', 'Sale Price of Property'),
            ('original_amount', 'Loan Amount'),
            ('down_payment', 'Down Payment'),
            ('deposit', 'Deposit (Earnest Money)'),
            ('seller_credit', 'Seller Credit'),
            ('cash_to_close', "Cash to Close"),
        ])
        section(display, 'Loan Terms', [
            ('interest_rate', 'Interest Rate'),
            ('monthly_payment', 'Monthly P&I Payment'),
            ('loan_term_years', 'Loan Term (years)'),
            ('loan_type', 'Loan Type'),
            ('apr', 'APR'),
        ])
        section(display, 'Operating Cost Estimates', [
            ('annual_property_tax', 'Annual Property Tax'),
            ('annual_insurance', 'Annual Homeowner\'s Insurance'),
            ('escrow_monthly', 'Monthly Escrow (taxes + insurance)'),
            ('hoa_annual', 'Annual HOA Dues (non-escrowed)'),
        ])

    if category == '1098':
        display = dict(data)
        full_addr = data.get('property_address', '')
        if data.get('property_city'):
            full_addr += f", {data['property_city']}"
        if data.get('property_state'):
            full_addr += f", {data['property_state']} {data.get('property_zip', '')}".rstrip()
        if full_addr:
            display['property_address'] = full_addr
        borrowers = '; '.join(
            data[k] for k in ('borrower_1', 'borrower_2', 'borrower_3', 'borrower_4')
            if data.get(k)
        )
        if borrowers:
            display['borrowers'] = borrowers

        section(display, 'Statement Information', [
            ('lender_name', 'Lender / Recipient'),
            ('borrowers', 'Borrowers'),
            ('account_number', 'Account Number'),
            ('tax_year', 'Calendar Year'),
            ('property_address', 'Property Securing Mortgage'),
        ])
        section(display, 'Mortgage Interest (Form 1098)', [
            ('mortgage_interest', 'Box 1 - Mortgage Interest Received'),
            ('current_balance', 'Box 2 - Outstanding Principal'),
            ('origination_date', 'Box 3 - Origination Date'),
            ('mortgage_insurance', 'Box 5 - Mortgage Insurance Premiums'),
            ('points_paid', 'Box 6 - Points Paid'),
            ('property_tax_amount', 'Box 10 - Real Estate Taxes Paid'),
        ])

    # Raw key/value schema — what the app imports
    schema_rows = [
        (k, f'{v:,.2f}' if isinstance(v, float) and 'rate' not in k else v)
        for k, v in data.items()
        if k != 'raw_text_preview' and v is not None
    ]
    if schema_rows:
        lines.append('## Recommended Extraction Schema for App Import')
        lines.append('')
        lines.extend(_md_table(schema_rows))

    return '\n'.join(lines)


def parse_document(filepath: str, category: str = 'auto') -> tuple[str, Dict[str, Any], str]:
    """Main entry: parse a document, returning (category, extracted data, markdown).

    With category='auto' the type is detected from the PDF text. The markdown
    is the internal structured representation generated before app import.
    """
    path = Path(filepath)
    ext = path.suffix.lower()

    raw_data = {}

    if ext == '.pdf':
        text = extract_pdf_text(filepath)
        if category == 'auto':
            category = detect_category(text)
        if category == 'mortgage_statement':
            raw_data = parse_mortgage_statement(text)
        elif category == 'tax_return':
            # Per-property Schedule E / Schedule A figures are stored separately
            # as TaxReturnEntry rows; the document itself just carries a clean
            # year/summary (the flat text parse mis-reads single totals).
            tr = parse_tax_return_properties(filepath)
            rentals = [p for p in tr.get('properties', []) if p['property_kind'] == 'rental']
            raw_data = {
                'tax_year': tr.get('tax_year'),
                'statement_year': tr.get('tax_year'),
                'property_count': len(tr.get('properties', [])),
                'rental_count': len(rentals),
                'total_rents_received': round(sum(p['rents_received'] for p in rentals), 2),
                'total_mortgage_interest': round(
                    sum(p['mortgage_interest'] for p in tr.get('properties', [])), 2),
                'period_type': 'yearly',
            }
        elif category == '1098':
            raw_data = parse_1098(text)
        elif category == '1099':
            raw_data = parse_1099(text)
        elif category == 'closing_statement':
            raw_data = parse_closing_statement(text)
        elif category == 'loan_disclosure':
            raw_data = parse_loan_disclosure(text)
        elif category == 'bank_statement':
            raw_data = parse_bank_statement(text)
        elif category == 'property_tax':
            raw_data = parse_property_tax(text)
        else:
            raw_data = {'raw_text_preview': text[:500]}

        # Add period type detection for parsed PDF documents
        if 'raw_text_preview' not in raw_data:
            raw_data['period_type'] = detect_period_type(raw_data)
    elif ext in ('.xlsx', '.xls'):
        if category == 'auto':
            category = 'other'
        raw_data = extract_excel_data(filepath)
    elif category == 'auto':
        category = 'other'

    markdown = to_markdown(category, raw_data) if isinstance(raw_data, dict) else ''
    return category, raw_data, markdown
